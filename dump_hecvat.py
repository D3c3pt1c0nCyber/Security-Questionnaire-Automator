import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
import re

FILE = r"C:\Users\joie.sayen\Downloads\HECVAT412_Template.xlsx"

# Load twice: once with data_only, once without (to see formulas)
wb_data = openpyxl.load_workbook(FILE, data_only=True)
wb_formula = openpyxl.load_workbook(FILE, data_only=False)

print("=" * 100)
print("HECVAT 4.1.2 TEMPLATE - COMPLETE STRUCTURE DUMP")
print("=" * 100)
print(f"\nAll sheet names: {wb_data.sheetnames}")
print(f"Total sheets: {len(wb_data.sheetnames)}")

TARGET_SHEETS = [
    "START HERE",
    "Organization",
    "Product",
    "Infrastructure",
    "IT Accessibility",
    "Case-Specific",
    "AI",
    "Privacy",
]

# Pattern for question IDs like XXXX-NN
qid_pattern = re.compile(r'^[A-Z]{3,5}-\d{1,3}$')

all_requ_questions = []

for sheet_name in TARGET_SHEETS:
    print("\n" + "=" * 100)
    print(f"SHEET: {sheet_name}")
    print("=" * 100)

    if sheet_name not in wb_data.sheetnames:
        print(f"  *** SHEET NOT FOUND ***")
        continue

    ws_data = wb_data[sheet_name]
    ws_formula = wb_formula[sheet_name]

    print(f"  Dimensions: {ws_data.dimensions}")
    print(f"  Max row: {ws_data.max_row}, Max col: {ws_data.max_column}")

    # First, print header rows (first 8 rows) to understand structure
    print(f"\n  --- HEADER ROWS (first 8 rows) ---")
    for row_idx in range(1, min(9, ws_data.max_row + 1)):
        cells = []
        for col_idx in range(1, min(ws_data.max_column + 1, 10)):
            val_data = ws_data.cell(row=row_idx, column=col_idx).value
            val_formula = ws_formula.cell(row=row_idx, column=col_idx).value
            if val_data is not None or val_formula is not None:
                if val_data != val_formula and val_formula is not None:
                    cells.append(f"Col{col_idx}: DATA={repr(val_data)[:80]} | FORMULA={repr(val_formula)[:80]}")
                else:
                    cells.append(f"Col{col_idx}: {repr(val_data)[:80]}")
        if cells:
            print(f"  Row {row_idx}: {' | '.join(cells)}")

    # Now find and dump all question IDs
    print(f"\n  --- ALL QUESTION ROWS ---")
    question_count = 0
    for row_idx in range(1, ws_data.max_row + 1):
        col_a_data = ws_data.cell(row=row_idx, column=1).value
        col_a_formula = ws_formula.cell(row=row_idx, column=1).value

        # Check if it looks like a question ID
        val_to_check = str(col_a_data) if col_a_data else ""
        if qid_pattern.match(val_to_check):
            question_count += 1
            qid = val_to_check

            # Column B - question text
            col_b_data = ws_data.cell(row=row_idx, column=2).value
            col_b_formula = ws_formula.cell(row=row_idx, column=2).value

            # Column C - answer
            col_c_data = ws_data.cell(row=row_idx, column=3).value
            col_c_formula = ws_formula.cell(row=row_idx, column=3).value

            # Column D
            col_d_data = ws_data.cell(row=row_idx, column=4).value
            col_d_formula = ws_formula.cell(row=row_idx, column=4).value

            # Column E (sometimes has extra info)
            col_e_data = ws_data.cell(row=row_idx, column=5).value
            col_e_formula = ws_formula.cell(row=row_idx, column=5).value

            print(f"\n  Row {row_idx} | {qid}")

            # Col B
            if col_b_data != col_b_formula and col_b_formula is not None:
                print(f"    B (question):  DATA={repr(col_b_data)[:120]}")
                print(f"    B (formula):   {repr(col_b_formula)[:120]}")
            else:
                print(f"    B (question):  {repr(col_b_data)[:150]}")

            # Col C
            if col_c_data != col_c_formula and col_c_formula is not None:
                print(f"    C (answer):    DATA={repr(col_c_data)[:120]}")
                print(f"    C (formula):   {repr(col_c_formula)[:120]}")
            elif col_c_data is not None:
                print(f"    C (answer):    {repr(col_c_data)[:150]}")
            else:
                print(f"    C (answer):    <empty>")

            # Col D
            if col_d_data != col_d_formula and col_d_formula is not None:
                print(f"    D:             DATA={repr(col_d_data)[:120]}")
                print(f"    D (formula):   {repr(col_d_formula)[:120]}")
            elif col_d_data is not None:
                print(f"    D:             {repr(col_d_data)[:150]}")

            # Col E
            if col_e_data is not None or (col_e_formula is not None and col_e_formula != col_e_data):
                if col_e_data != col_e_formula and col_e_formula is not None:
                    print(f"    E:             DATA={repr(col_e_data)[:120]}")
                    print(f"    E (formula):   {repr(col_e_formula)[:120]}")
                elif col_e_data is not None:
                    print(f"    E:             {repr(col_e_data)[:150]}")

            # Track REQU questions
            if qid.startswith("REQU-"):
                all_requ_questions.append({
                    "qid": qid,
                    "row": row_idx,
                    "sheet": sheet_name,
                    "question": col_b_data or col_b_formula,
                    "answer_data": col_c_data,
                    "answer_formula": col_c_formula,
                    "col_d": col_d_data,
                })

    print(f"\n  >>> Total questions in '{sheet_name}': {question_count}")

# Also scan for any non-target sheets that might have questions
print("\n" + "=" * 100)
print("SCANNING OTHER SHEETS FOR QUESTION IDs")
print("=" * 100)
for sn in wb_data.sheetnames:
    if sn not in TARGET_SHEETS:
        ws = wb_data[sn]
        count = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val and qid_pattern.match(str(val)):
                count += 1
                if count <= 5:
                    print(f"  [{sn}] Row {row_idx}: {val} | B={repr(ws.cell(row=row_idx, column=2).value)[:80]}")
        if count > 5:
            print(f"  [{sn}] ... and {count - 5} more questions (total: {count})")
        elif count == 0:
            print(f"  [{sn}] No question IDs found (max_row={ws.max_row}, max_col={ws.max_column})")

# REQU questions summary
print("\n" + "=" * 100)
print("REQU QUESTIONS SUMMARY (Control which Case-Specific sections are required)")
print("=" * 100)
if all_requ_questions:
    for rq in all_requ_questions:
        print(f"\n  {rq['qid']} (Sheet: {rq['sheet']}, Row: {rq['row']})")
        print(f"    Question: {repr(rq['question'])[:150]}")
        print(f"    Answer (data):    {repr(rq['answer_data'])[:120]}")
        if rq['answer_formula'] != rq['answer_data']:
            print(f"    Answer (formula): {repr(rq['answer_formula'])[:120]}")
        if rq['col_d']:
            print(f"    Col D:            {repr(rq['col_d'])[:120]}")
else:
    print("  No REQU-XX questions found!")
    print("  Searching for 'REQU' anywhere in the workbook...")
    for sn in wb_data.sheetnames:
        ws_d = wb_data[sn]
        ws_f = wb_formula[sn]
        for row_idx in range(1, ws_d.max_row + 1):
            for col_idx in range(1, min(ws_d.max_column + 1, 8)):
                vd = ws_d.cell(row=row_idx, column=col_idx).value
                vf = ws_f.cell(row=row_idx, column=col_idx).value
                for v in [vd, vf]:
                    if v and "REQU" in str(v).upper():
                        print(f"    [{sn}] Row {row_idx}, Col {col_idx}: DATA={repr(vd)[:100]} FORMULA={repr(vf)[:100]}")
                        break

# Data validation / dropdowns
print("\n" + "=" * 100)
print("DATA VALIDATIONS (Dropdowns) IN TARGET SHEETS")
print("=" * 100)
for sheet_name in TARGET_SHEETS:
    if sheet_name not in wb_data.sheetnames:
        continue
    ws = wb_data[sheet_name]
    if ws.data_validations and ws.data_validations.dataValidation:
        print(f"\n  [{sheet_name}] {len(ws.data_validations.dataValidation)} validation(s):")
        for dv in ws.data_validations.dataValidation:
            print(f"    Range: {dv.sqref} | Type: {dv.type} | Formula1: {repr(dv.formula1)[:100]} | Allow blank: {dv.allow_blank}")
    else:
        print(f"\n  [{sheet_name}] No data validations found")

# Grand summary
print("\n" + "=" * 100)
print("GRAND SUMMARY")
print("=" * 100)
total = 0
for sheet_name in TARGET_SHEETS:
    if sheet_name not in wb_data.sheetnames:
        print(f"  {sheet_name}: NOT FOUND")
        continue
    ws = wb_data[sheet_name]
    count = 0
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val and qid_pattern.match(str(val)):
            count += 1
    total += count
    print(f"  {sheet_name:20s}: {count} questions")
print(f"  {'TOTAL':20s}: {total} questions")

wb_data.close()
wb_formula.close()
print("\nDone.")
